#!/usr/bin/env python3
"""
Generate readable Excel files for each vendor from catalog manifests.
"""

import json
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def load_vendor_manifest(manifest_path):
    """Load vendor manifest JSON file."""
    with open(manifest_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def create_vendor_excel(vendor_data, output_path):
    """Create Excel file for a vendor with all integrations and use cases."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Integrations"

    # Define headers
    headers = [
        "Vendor Name",
        "Vendor Description",
        "Vendor Category",
        "Vendor Tags",
        "Vendor Homepage",
        "Vendor Logo URL",
        "Registration Required",
        "Subscription Required",
        "Free Subscription",
        "Type",
        "Item Name",
        "Item Description",
        "Version",
        "Data Types",
        "Item Tags",
        "File Path",
        "URL"
    ]

    # Write headers with styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Extract vendor-level info
    vendor_name = vendor_data.get("name", "")
    vendor_description = vendor_data.get("description", "")
    vendor_category = vendor_data.get("category", "")
    vendor_tags = ", ".join(vendor_data.get("tags", []))
    vendor_homepage = vendor_data.get("homepage", "")

    # Handle logo URL - can be direct or nested in light/dark variants
    logo = vendor_data.get("logo", {})
    if isinstance(logo, dict):
        if "url" in logo:
            vendor_logo_url = logo.get("url", "")
        elif "light" in logo:
            vendor_logo_url = logo.get("light", {}).get("url", "")
        else:
            vendor_logo_url = ""
    else:
        vendor_logo_url = ""

    vendor_registration_required = vendor_data.get("registration_required", "")
    vendor_subscription_required = vendor_data.get("subscription_required", "")
    vendor_free_subscription = vendor_data.get("free_subscription", "")

    row_num = 2

    # Add analyzers
    integrations = vendor_data.get("integrations", {})
    for analyzer in integrations.get("analyzers", []):
        ws.cell(row=row_num, column=1).value = vendor_name
        ws.cell(row=row_num, column=2).value = vendor_description
        ws.cell(row=row_num, column=3).value = vendor_category
        ws.cell(row=row_num, column=4).value = vendor_tags
        ws.cell(row=row_num, column=5).value = vendor_homepage
        ws.cell(row=row_num, column=6).value = vendor_logo_url
        ws.cell(row=row_num, column=7).value = vendor_registration_required
        ws.cell(row=row_num, column=8).value = vendor_subscription_required
        ws.cell(row=row_num, column=9).value = vendor_free_subscription
        ws.cell(row=row_num, column=10).value = "Analyzer"
        ws.cell(row=row_num, column=11).value = analyzer.get("name", "")
        ws.cell(row=row_num, column=12).value = analyzer.get("description", "")
        ws.cell(row=row_num, column=13).value = analyzer.get("version", "")
        ws.cell(row=row_num, column=14).value = ", ".join(analyzer.get("dataTypes", []))
        ws.cell(row=row_num, column=15).value = ""
        ws.cell(row=row_num, column=16).value = analyzer.get("file", "")
        ws.cell(row=row_num, column=17).value = analyzer.get("url", "")
        row_num += 1

    # Add responders
    for responder in integrations.get("responders", []):
        ws.cell(row=row_num, column=1).value = vendor_name
        ws.cell(row=row_num, column=2).value = vendor_description
        ws.cell(row=row_num, column=3).value = vendor_category
        ws.cell(row=row_num, column=4).value = vendor_tags
        ws.cell(row=row_num, column=5).value = vendor_homepage
        ws.cell(row=row_num, column=6).value = vendor_logo_url
        ws.cell(row=row_num, column=7).value = vendor_registration_required
        ws.cell(row=row_num, column=8).value = vendor_subscription_required
        ws.cell(row=row_num, column=9).value = vendor_free_subscription
        ws.cell(row=row_num, column=10).value = "Responder"
        ws.cell(row=row_num, column=11).value = responder.get("name", "")
        ws.cell(row=row_num, column=12).value = responder.get("description", "")
        ws.cell(row=row_num, column=13).value = responder.get("version", "")
        ws.cell(row=row_num, column=14).value = ", ".join(responder.get("dataTypes", []))
        ws.cell(row=row_num, column=15).value = ""
        ws.cell(row=row_num, column=16).value = responder.get("file", "")
        ws.cell(row=row_num, column=17).value = responder.get("url", "")
        row_num += 1

    # Add functions
    for function in integrations.get("functions", []):
        ws.cell(row=row_num, column=1).value = vendor_name
        ws.cell(row=row_num, column=2).value = vendor_description
        ws.cell(row=row_num, column=3).value = vendor_category
        ws.cell(row=row_num, column=4).value = vendor_tags
        ws.cell(row=row_num, column=5).value = vendor_homepage
        ws.cell(row=row_num, column=6).value = vendor_logo_url
        ws.cell(row=row_num, column=7).value = vendor_registration_required
        ws.cell(row=row_num, column=8).value = vendor_subscription_required
        ws.cell(row=row_num, column=9).value = vendor_free_subscription
        ws.cell(row=row_num, column=10).value = "Function"
        ws.cell(row=row_num, column=11).value = function.get("name", "")
        ws.cell(row=row_num, column=12).value = function.get("description", "")
        ws.cell(row=row_num, column=13).value = function.get("version", "")
        ws.cell(row=row_num, column=14).value = ""
        ws.cell(row=row_num, column=15).value = ""
        ws.cell(row=row_num, column=16).value = function.get("file", "")
        ws.cell(row=row_num, column=17).value = function.get("url", "")
        row_num += 1

    # Add use cases
    for use_case in vendor_data.get("useCases", []):
        ws.cell(row=row_num, column=1).value = vendor_name
        ws.cell(row=row_num, column=2).value = vendor_description
        ws.cell(row=row_num, column=3).value = vendor_category
        ws.cell(row=row_num, column=4).value = vendor_tags
        ws.cell(row=row_num, column=5).value = vendor_homepage
        ws.cell(row=row_num, column=6).value = vendor_logo_url
        ws.cell(row=row_num, column=7).value = vendor_registration_required
        ws.cell(row=row_num, column=8).value = vendor_subscription_required
        ws.cell(row=row_num, column=9).value = vendor_free_subscription
        ws.cell(row=row_num, column=10).value = "Use Case"
        ws.cell(row=row_num, column=11).value = use_case.get("name", "")
        ws.cell(row=row_num, column=12).value = use_case.get("description", "")
        ws.cell(row=row_num, column=13).value = ""
        ws.cell(row=row_num, column=14).value = ""
        ws.cell(row=row_num, column=15).value = ", ".join(use_case.get("tags", []))

        # Get documentation file and URL
        doc = use_case.get("documentation", {})
        ws.cell(row=row_num, column=16).value = doc.get("file", "")
        ws.cell(row=row_num, column=17).value = doc.get("url", "")
        row_num += 1

    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0

        for row in ws[column_letter]:
            try:
                if row.value:
                    max_length = max(max_length, len(str(row.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Enable text wrapping for description columns
    for row in range(2, row_num):
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=12).alignment = Alignment(wrap_text=True, vertical="top")

    # Save workbook
    wb.save(output_path)
    print(f"Generated: {output_path}")


def main():
    """Main function to process all vendors."""
    catalogs_dir = Path("catalogs")
    output_dir = Path("vendor-excel-exports")

    # Create output directory if it doesn't exist
    output_dir.mkdir(exist_ok=True)

    # Find all vendor manifest files (exclude integration-manifest.json)
    manifest_files = []
    for vendor_dir in catalogs_dir.iterdir():
        if vendor_dir.is_dir():
            manifest_path = vendor_dir / "manifest.json"
            if manifest_path.exists():
                manifest_files.append(manifest_path)

    if not manifest_files:
        print("No vendor manifests found in catalogs directory")
        return

    print(f"Found {len(manifest_files)} vendor manifest(s)")

    # Process each vendor
    for manifest_path in manifest_files:
        try:
            vendor_data = load_vendor_manifest(manifest_path)
            vendor_id = vendor_data.get("id", "unknown")

            output_file = output_dir / f"{vendor_id}.xlsx"
            create_vendor_excel(vendor_data, output_file)

        except Exception as e:
            print(f"Error processing {manifest_path}: {e}")

    print(f"\nAll Excel files generated in '{output_dir}' directory")


if __name__ == "__main__":
    main()
